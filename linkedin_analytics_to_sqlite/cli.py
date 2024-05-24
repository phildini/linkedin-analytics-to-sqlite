import click
from openpyxl import load_workbook
from sqlite_utils import Database


@click.group()
@click.version_option()
def cli():
    """
    Save data from a LinkedIn Analytics export xlsx to a SQLite database.
    """


@cli.command()
@click.argument(
    "xlsx_path",
    type=click.Path(file_okay=True, dir_okay=False, allow_dash=False),
    required=True,
)
@click.option(
    "-d",
    "--db-path",
    type=click.Path(file_okay=True, dir_okay=False, allow_dash=False),
    default="linkedin_analytics.db",
)
def process(xlsx_path, db_path):
    db = Database(db_path)
    db["engagement"].create(
        {
            "date": str,
            "impressions": str,
            "engagements": str,
        },
        pk="date",
        if_not_exists=True,
    )
    db["followers"].create(
        {
            "date": str,
            "new_followers": str,
        },
        pk="date",
        if_not_exists=True,
    )
    db["demographics"].create(
        {
            "id": str,
            "demographic_type": str,
            "value": str,
            "percentage": float,
            "start_date": str,
            "end_date": str,
        },
        pk="id",
        if_not_exists=True,
    )

    db["top_posts"].create(
        {
            "post_url": str,
            "publish_date": str,
            "engagements": int,
            "impressions": int,
        },
        pk="post_url",
        if_not_exists=True,
    )

    # This sucks and I'm sure there's a better way.
    dates = xlsx_path.split("/")[-1].split("_")[1:3]

    wb = load_workbook(
        filename=xlsx_path,
        read_only=True,
    )
    db["engagement"].upsert_all(parse_engagement(wb["ENGAGEMENT"]), pk="date")
    db["followers"].upsert_all(parse_followers(wb["FOLLOWERS"]), pk="date")
    db["demographics"].upsert_all(
        parse_demographics(wb["DEMOGRAPHICS"], dates[0], dates[1]), pk="id"
    )
    db["top_posts"].upsert_all(parse_top_posts(wb["TOP POSTS"]), pk="post_url")
    # print(parse_engagement(wb["DEMOGRAPHICS"]))


def parse_engagement(sheet):
    engagement_dates = []
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=1000, values_only=True):
        engagement_dates.append(
            {
                "date": row[0],
                "impressions": row[1],
                "engagements": row[2],
            }
        )
    return engagement_dates


def parse_followers(sheet):
    follower_dates = []
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=1000, values_only=True):
        follower_dates.append(
            {
                "date": row[0],
                "new_followers": row[1],
            }
        )
    return follower_dates


def parse_demographics(sheet, start_date, end_date):
    demographics = []
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=1000, values_only=True):
        id = f"{row[0]}_{row[1]}_{start_date}_{end_date}"
        demographics.append(
            {
                "demographic_type": row[0],
                "value": row[1],
                "percentage": row[2],
                "start_date": start_date,
                "end_date": end_date,
                "id": id,
            }
        )
    return demographics


def parse_top_posts(sheet):
    top_posts = {}

    # Get impressions
    for row in sheet.iter_rows(
        min_row=4, min_col=5, max_col=7, max_row=1000, values_only=True
    ):
        top_posts[row[0]] = {
            "post_url": row[0],
            "publish_date": row[1],
            "impressions": row[2],
            "engagements": 0,
        }

    # Get engagements
    for row in sheet.iter_rows(min_row=4, max_col=3, max_row=1000, values_only=True):
        if row[0]:
            top_posts[row[0]]["engagements"] = row[2]

    return top_posts.values()
